from openpyxl import Workbook, load_workbook
from datetime import datetime
#-------------------------------------------------------------
#SETUP
#-------------------------------------------------------------
wb1 = load_workbook('MasterPriceList.xlsx')
wb2 = load_workbook('TemplatePriceList.xlsx')

ws1 = wb1.active
ws2 = wb2.active
#-------------------------------------------------------------
#FUNCTIONS
#-------------------------------------------------------------
def getColNum(str):
    if str == "Prefix":
        return 1
    elif str == "Product Code":
        return 2
    elif str == "Product Description":
        return 3
    elif str == "UPC":
        return 4
    elif str == "List Price":
        return 5
    elif str == "Net Cost":
        return 6
    elif str == "Product Line":
        return 7
    elif str == "Price Group":
        return 8
    elif str == "Placeholder 1":
        return 9
    elif str == "Placeholder 2":
        return 10
    elif str == "Placeholder 3":
        return 11
    elif str == "Last Updated":
        return 12
    else:
        return 5


def ingestToList(ws):
    sheet_cells = []
    for rows in ws.iter_rows():
        row_cells = []
        for cell in rows:
            row_cells.append(cell.value)
        sheet_cells.append(row_cells)
    return sheet_cells


def ingestToDict(ws):
    sheet_cells = {}
    rowNum = 1
    for row in ws.iter_rows():
        pCode = row[1].value
        sheet_cells[pCode] = rowNum
        rowNum += 1
    return sheet_cells


def printList(pList):
    for i in range(len(pList)):
        for j in range(len(pList[i])):
            print(pList[i][j], end=' ')
        print("\n")


def merge(master, template, masterCol):
    for row in template:
        if row[1] in master:
            masterIdx = master[row[1]]
            ws1.cell(row=masterIdx, column=masterCol, value=row[4])
            now = datetime.now()
            ws1.cell(row=masterIdx, column=12, value=now)
        else:
            newRow = []
            now = datetime.now()
            for v in row:
                newRow.append(v)
            newRow.append(now)
            ws1.append(newRow)
#-------------------------------------------------------------
#IMPLEMENTATION
#-------------------------------------------------------------
templateList = ingestToList(ws2)
masterDict = ingestToDict(ws1)
masterCol = getColNum("List Price")
merge(masterDict, templateList, masterCol)
#-------------------------------------------------------------
#CLOSING
#-------------------------------------------------------------
wb1.save('MasterPriceList.xlsx')
wb2.save('TemplatePriceList.xlsx')