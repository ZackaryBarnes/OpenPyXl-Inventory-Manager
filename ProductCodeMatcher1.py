from openpyxl import Workbook, load_workbook
from datetime import datetime
#-------------------------------------------------------------
#SETUP
#-------------------------------------------------------------
wb1 = load_workbook('MasterPriceList.xlsx')
wb2 = load_workbook('TemplatePriceList.xlsx')

ws1 = wb1.active
ws2 = wb2.active
#-------------------------------------------------------------
#FUNCTIONS
#-------------------------------------------------------------
def createRow(row):
    row_cells = []
    newPCode = f"{row[0].value}-{row[1].value}"
    row_cells.append(newPCode)
    for cell in range(2, len(row)):
        row_cells.append(row[cell].value)
    now = datetime.now()
    row_cells.append(now)
    return row_cells
    

def ingestToList(ws):
    sheet_cells = []
    for row in ws.iter_rows(min_row=2):
        newRow = createRow(row)
        sheet_cells.append(newRow)
    return sheet_cells


def ingestToDict(ws):
    sheet_cells = {}
    rowNum = 2
    for row in ws.iter_rows(min_row=2):
        pCode = row[0].value
        sheet_cells[pCode] = rowNum
        rowNum += 1
    return sheet_cells


def printList(pList):
    for i in range(len(pList)):
        for j in range(len(pList[i])):
            print(pList[i][j], end=' ')
        print("\n")


def mergeSheets(master, template):
    for row in template:
        if row[0] in master:
            masterIdx = master[row[0]]
            ws1.cell(row=masterIdx, column=4, value=row[3])
            now = datetime.now()
            ws1.cell(row=masterIdx, column=11, value=now)
        else:
            ws1.append(row)
#-------------------------------------------------------------
#IMPLEMENTATION
#-------------------------------------------------------------
templateList = ingestToList(ws2)
masterDict = ingestToDict(ws1)
mergeSheets(masterDict, templateList)
#-------------------------------------------------------------
#CLOSING
#-------------------------------------------------------------
wb1.save('MasterPriceList.xlsx')
wb2.save('TemplatePriceList.xlsx')